import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()

# ============================================================
# Style definitions
# ============================================================
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
red_font = Font(color="9C0006")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(color="006100")

def style_header_row(ws, row=1, max_col=None):
    if max_col is None:
        max_col = ws.max_column
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

def style_data_area(ws, start_row=2, max_row=None, max_col=None):
    if max_row is None:
        max_row = ws.max_row
    if max_col is None:
        max_col = ws.max_column
    for row in range(start_row, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

def auto_width(ws, max_col=None, min_width=10, max_width=40):
    if max_col is None:
        max_col = ws.max_column
    for col in range(1, max_col + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len

def add_data_validation(ws, col_letter, formula1, start_row=2, end_row=1000):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.error = "Please select a valid option"
    dv.errorTitle = "Invalid Entry"
    dv.prompt = "Select from dropdown"
    dv.promptTitle = "Allowed Values"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")

# ============================================================
# TAB 1: Enquiry Log
# ============================================================
ws1 = wb.active
ws1.title = "Enquiry Log"

enquiry_headers = ["Date/Time", "Channel", "Customer Name", "Phone", "Product/Enquiry", 
                   "Urgency", "Status", "Quote Sent?", "Quote Price", "Notes"]
for col, h in enumerate(enquiry_headers, 1):
    ws1.cell(row=1, column=col, value=h)
style_header_row(ws1, max_col=len(enquiry_headers))

# Sample data
enquiry_data = [
    ["14/07/26 06:42", "Jiji", "Kwame A.", "024XXXXXXX", "Makita Drill 18V", "Hot", "Quoted", "Yes", "GHS 1,250", "Replied in 3 min"],
    ["14/07/26 07:15", "WhatsApp", "Ama K.", "020XXXXXXX", "Bosch Grinder", "Warm", "Checking stock", "No", "—", "Checking Zobaze now"]
]
for row_idx, row_data in enumerate(enquiry_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws1.cell(row=row_idx, column=col_idx, value=val)

# Data validations
add_data_validation(ws1, "B", '"Call,WhatsApp,Jiji,Walk-in,Referral"')
add_data_validation(ws1, "F", '"Hot,Warm,Cold"')
add_data_validation(ws1, "G", '"New,Checking Stock,Quoted,Sourcing,Converted,Lost,Follow-up"')
add_data_validation(ws1, "H", '"Yes,No"')

style_data_area(ws1, max_row=1000, max_col=10)
auto_width(ws1, max_col=10)

# ============================================================
# TAB 2: Orders Log
# ============================================================
ws2 = wb.create_sheet("Orders Log")
orders_headers = ["Order ID", "Date", "Customer", "Phone", "SKU", "Qty", "Unit Price", 
                  "Total", "Payment Ref", "Payment Method", "Status", "Rider", 
                  "Dispatch Time", "Notes"]
for col, h in enumerate(orders_headers, 1):
    ws2.cell(row=1, column=col, value=h)
style_header_row(ws2, max_col=len(orders_headers))

orders_data = [
    ["ORD-20260714-001", "14/07/26", "Kwame A.", "024XXXXXXX", "MAK-DRL-18V", 1, 1250, 1250, 
     "MTN-240714-001", "MoMo", "Dispatched", "Yaw K.", "14/07 10:15", "Delivered 11:30"]
]
for row_idx, row_data in enumerate(orders_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws2.cell(row=row_idx, column=col_idx, value=val)

add_data_validation(ws2, "K", '"Confirmed,Packed,Dispatched,Delivered,Completed,Returned,Cancelled"')
add_data_validation(ws2, "L", '"MoMo,Cash,Bank Transfer"')

style_data_area(ws2, max_row=1000, max_col=14)
auto_width(ws2, max_col=14)

# ============================================================
# TAB 3: Sales Log
# ============================================================
ws3 = wb.create_sheet("Sales Log")
sales_headers = ["Date", "Order ID", "Channel", "Customer", "SKU", "Qty", 
                 "Unit Price", "Total", "Payment Method", "Rider", "Status"]
for col, h in enumerate(sales_headers, 1):
    ws3.cell(row=1, column=col, value=h)
style_header_row(ws3, max_col=len(sales_headers))

sales_data = [
    ["14/07/26", "ORD-20260714-001", "Jiji", "Kwame A.", "MAK-DRL-18V", 1, 1250, 1250, "MoMo", "Yaw K.", "Completed"]
]
for row_idx, row_data in enumerate(sales_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws3.cell(row=row_idx, column=col_idx, value=val)

# Reconciliation section (rows below data)
recon_start = 100
ws3.cell(row=recon_start, column=1, value="End-of-Day Reconciliation (17:00)").font = Font(bold=True, size=12)
ws3.cell(row=recon_start+1, column=1, value="Zobaze Total Sales (GHS):")
ws3.cell(row=recon_start+1, column=2).border = thin_border
ws3.cell(row=recon_start+2, column=1, value="Tracker Total Sales (GHS):")
ws3.cell(row=recon_start+2, column=2, value="=SUM(H2:H1000)").border = thin_border
ws3.cell(row=recon_start+3, column=1, value="Variance:")
ws3.cell(row=recon_start+3, column=2, value=f"=B{recon_start+2}-B{recon_start+1}").border = thin_border
ws3.cell(row=recon_start+4, column=1, value="Notes on any variance:")

add_data_validation(ws3, "K", '"Confirmed,Packed,Dispatched,Delivered,Completed,Returned,Cancelled"')
add_data_validation(ws3, "J", '"MoMo,Cash,Bank Transfer"')

style_data_area(ws3, max_row=1000, max_col=11)
auto_width(ws3, max_col=11)

# ============================================================
# TAB 4: Dispatch Log
# ============================================================
ws4 = wb.create_sheet("Dispatch Log")
dispatch_headers = ["Order ID", "Date", "Time Dispatched", "Rider Name", "Rider Phone", 
                    "Rider Plate", "Tracking Link", "Items Dispatched", "Packed By", "Photo Sent?"]
for col, h in enumerate(dispatch_headers, 1):
    ws4.cell(row=1, column=col, value=h)
style_header_row(ws4, max_col=len(dispatch_headers))

dispatch_data = [
    ["ORD-20260714-001", "14/07/26", "10:15", "Yaw K.", "024XXXXXXX", "GT-XXXX-24", 
     "[Yango link]", "MAK-DRL-18V x1", "John", "Yes"]
]
for row_idx, row_data in enumerate(dispatch_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws4.cell(row=row_idx, column=col_idx, value=val)

add_data_validation(ws4, "J", '"Yes,No"')
style_data_area(ws4, max_row=1000, max_col=10)
auto_width(ws4, max_col=10)

# ============================================================
# TAB 5: Deliveries Log
# ============================================================
ws5 = wb.create_sheet("Deliveries Log")
deliveries_headers = ["Order ID", "Date Delivered", "Time Delivered", "Rider", 
                      "Customer Confirmed?", "Confirmation Time", "Issue?", "Issue Details", "Status"]
for col, h in enumerate(deliveries_headers, 1):
    ws5.cell(row=1, column=col, value=h)
style_header_row(ws5, max_col=len(deliveries_headers))

deliveries_data = [
    ["ORD-20260714-001", "14/07/26", "11:30", "Yaw K.", "Yes", "11:35", "No", "—", "Completed"]
]
for row_idx, row_data in enumerate(deliveries_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws5.cell(row=row_idx, column=col_idx, value=val)

# Rule note
ws5.cell(row=100, column=1, value="Rule: Call customer within 30 min of drop-off. Log confirmation time.").font = Font(italic=True, color="666666")

add_data_validation(ws5, "E", '"Yes,No"')
add_data_validation(ws5, "G", '"Yes,No"')
add_data_validation(ws5, "I", '"Completed,Partial,Failed,Rescheduled"')
style_data_area(ws5, max_row=1000, max_col=9)
auto_width(ws5, max_col=9)

# ============================================================
# TAB 6: Stock Availability
# ============================================================
ws6 = wb.create_sheet("Stock Availability")
stock_headers = ["SKU", "Product Name", "Category", "Zobaze Qty", "Physical Count", 
                 "Variance", "Min Stock", "Reorder?", "Location (Shelf)", "Last Updated"]
for col, h in enumerate(stock_headers, 1):
    ws6.cell(row=1, column=col, value=h)
style_header_row(ws6, max_col=len(stock_headers))

stock_data = [
    ["MAK-DRL-18V", "Makita 18V Drill", "Power Tools", 3, 3, 0, 2, "No", "A-1-Top", "14/07 07:05"],
    ["BOS-GRD-900", "Bosch 900W Grinder", "Power Tools", 0, 0, 0, 1, "YES", "—", "14/07 07:05"]
]
for row_idx, row_data in enumerate(stock_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws6.cell(row=row_idx, column=col_idx, value=val)

# Variance formula (F = D - E)
for row in range(2, 1001):
    ws6.cell(row=row, column=6).value = f"=D{row}-E{row}"

# Reorder? formula (H = IF(E<=G, "YES", "No"))
for row in range(2, 1001):
    ws6.cell(row=row, column=8).value = f'=IF(E{row}<=G{row},"YES","No")'

add_data_validation(ws6, "H", '"YES,No"')

# Conditional formatting: Red if Variance ≠ 0 (col F), Red if Reorder? = YES (col H)
ws6.conditional_formatting.add("F2:F1000", CellIsRule(operator="notEqual", formula=["0"], fill=red_fill, font=red_font))
ws6.conditional_formatting.add("H2:H1000", CellIsRule(operator="equal", formula=['"YES"'], fill=red_fill, font=red_font))

style_data_area(ws6, max_row=1000, max_col=10)
auto_width(ws6, max_col=10)

# ============================================================
# TAB 7: Sourcing Log
# ============================================================
ws7 = wb.create_sheet("Sourcing Log")
sourcing_headers = ["Date", "Customer", "Phone", "SKU/Description", "Urgency", 
                    "Quote 1 (Source/Price)", "Quote 2 (Source/Price)", "Quote 3 (Source/Price)", 
                    "H Decision", "Final Price", "Status"]
for col, h in enumerate(sourcing_headers, 1):
    ws7.cell(row=1, column=col, value=h)
style_header_row(ws7, max_col=len(sourcing_headers))

sourcing_data = [
    ["14/07/26", "Kwame A.", "024XXXXXXX", "DeWalt 20V Impact", "High", 
     "UK eBay / GHS 1,850", "Direct / GHS 1,650", "Jiji Comp / GHS 2,100", 
     "Pending", "—", "Sourcing"]
]
for row_idx, row_data in enumerate(sourcing_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws7.cell(row=row_idx, column=col_idx, value=val)

add_data_validation(ws7, "E", '"High,Medium,Low"')
add_data_validation(ws7, "I", '"Approved,Rejected,Pending"')
add_data_validation(ws7, "K", '"Sourcing,Quoted,Approved,Ordered,Received,Cancelled"')

# SLA note
ws7.cell(row=100, column=1, value="SLA: 3 quotes in WhatsApp group within 24h of logging.").font = Font(italic=True, color="666666")

style_data_area(ws7, max_row=1000, max_col=11)
auto_width(ws7, max_col=11)

# ============================================================
# TAB 8: Jiji Stats
# ============================================================
ws8 = wb.create_sheet("Jiji Stats")
jiji_headers = ["Date", "Priority Listings Active", "Total Views", "Total Chats", 
                "Avg Response Time (min)", "Leads Generated", "Converted to Sale", 
                "Price Adjustments", "Expired Listings Relisted"]
for col, h in enumerate(jiji_headers, 1):
    ws8.cell(row=1, column=col, value=h)
style_header_row(ws8, max_col=len(jiji_headers))

jiji_data = [
    ["14/07/26", "20/20", 1240, 47, 3.2, 12, 3, "2 (MAK-DRL-18V -2%)", 0]
]
for row_idx, row_data in enumerate(jiji_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws8.cell(row=row_idx, column=col_idx, value=val)

style_data_area(ws8, max_row=1000, max_col=9)
auto_width(ws8, max_col=9)

# ============================================================
# TAB 9: Write-Off Log
# ============================================================
ws9 = wb.create_sheet("Write-Off Log")
writeoff_headers = ["Date", "SKU", "Product", "Qty", "Reason", "Value (GHS)", 
                    "Photo?", "Approved By", "Disposed?"]
for col, h in enumerate(writeoff_headers, 1):
    ws9.cell(row=1, column=col, value=h)
style_header_row(ws9, max_col=len(writeoff_headers))

writeoff_data = [
    ["14/07/26", "BOS-GRD-900", "Bosch Grinder", 1, "Damaged in transit", 450, "Yes", "H", "Yes"]
]
for row_idx, row_data in enumerate(writeoff_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws9.cell(row=row_idx, column=col_idx, value=val)

add_data_validation(ws9, "G", '"Yes,No"')
add_data_validation(ws9, "I", '"Yes,No"')
style_data_area(ws9, max_row=1000, max_col=9)
auto_width(ws9, max_col=9)

# ============================================================
# TAB 10: Daily Summary
# ============================================================
ws10 = wb.create_sheet("Daily Summary")
summary_headers = ["Metric", "Today", "This Week", "Target", "Variance"]
for col, h in enumerate(summary_headers, 1):
    ws10.cell(row=1, column=col, value=h)
style_header_row(ws10, max_col=len(summary_headers))

summary_data = [
    ["Enquiries Logged", "", "", "—", "—"],
    ["Enquiries → Quotes", "", "", "80%", ""],
    ["Conversion Rate", "", "", "25%", ""],
    ["Orders Processed", "", "", "—", "—"],
    ["Revenue (GHS)", "", "", "—", "—"],
    ["Dispatched", "", "", "—", "—"],
    ["Delivered & Confirmed", "", "", "100%", ""],
    ["Avg Dispatch→Delivery (hrs)", "", "", "<4", ""],
    ["Jiji Avg Response (min)", "", "", "<5", ""],
    ["Stock Variance (Top 50)", "", "", "<2%", ""],
    ["Sourcing Requests", "", "", "—", "—"],
    ["Sourcing Completed (24h)", "", "", "100%", ""]
]
for row_idx, row_data in enumerate(summary_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws10.cell(row=row_idx, column=col_idx, value=val)

# Formula hints
formula_start = 20
ws10.cell(row=formula_start, column=1, value="FORMULAS / AUTOMATION HINTS").font = Font(bold=True, size=12)
formulas = [
    "Order ID: =\"ORD-\"&TEXT(TODAY(),\"YYYYMMDD\")&\"-\"&TEXT(ROW()-1,\"000\")",
    "Daily Revenue: =SUMIFS('Sales Log'!H:H, 'Sales Log'!A:A, TODAY())",
    "Conversion Rate: =COUNTIF('Enquiry Log'!F:F,\"Converted\")/COUNTA('Enquiry Log'!A2:A)",
    "Stock Variance %: =ABS(SUM('Stock Availability'!D:D)-SUM('Stock Availability'!E:E))/SUM('Stock Availability'!D:D)",
    "Jiji Response Avg: =AVERAGE('Jiji Stats'!E:E)"
]
for i, f in enumerate(formulas):
    ws10.cell(row=formula_start+1+i, column=1, value=f).font = Font(size=10, color="666666")

# Sharing & Permissions
share_start = formula_start + 7
ws10.cell(row=share_start, column=1, value="SHARING & PERMISSIONS").font = Font(bold=True, size=12)
share_data = [
    ["User", "Access"],
    ["H (Owner)", "View + Comment"],
    ["John", "Edit"],
    ["(Optional) Accountant", "View (Sales Log only)"]
]
for row_idx, row_data in enumerate(share_data, share_start+1):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws10.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == share_start+1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        cell.border = thin_border

# Notes section
notes_start = share_start + 6
ws10.cell(row=notes_start, column=1, value="Notes / Issues / Escalations:").font = Font(bold=True)
ws10.cell(row=notes_start+1, column=1, value="").border = thin_border
ws10.cell(row=notes_start+2, column=1, value="").border = thin_border

ws10.cell(row=notes_start+4, column=1, value="Sent to H (WhatsApp) at:").font = Font(bold=True)

style_data_area(ws10, max_row=50, max_col=5)
auto_width(ws10, max_col=5, min_width=15, max_width=50)

# ============================================================
# Save
# ============================================================
output_path = r"C:\Users\User\.hermes\workspace\Vault\jobs\2REAL_DAILY_TRACKER_TEMPLATE.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Sheets: {wb.sheetnames}")