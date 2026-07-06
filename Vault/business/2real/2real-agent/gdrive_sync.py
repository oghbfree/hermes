#!/usr/bin/env python3
"""
2Real Agent — Google Drive Inventory Auto-Sync
Checks the zobaze inventory folder for updated files and syncs to agent database.
"""
import os
import json
import shutil
from datetime import datetime
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import openpyxl

# Configuration
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
WORKSPACE = os.path.dirname(os.path.abspath(__file__))
CREDS_FILE = os.path.join(WORKSPACE, 'gdrive_credentials.json')
TOKEN_FILE = os.path.join(WORKSPACE, 'gdrive_token.json')
INVENTORY_FOLDER_ID = '1ENfsztK1K7aw6TrtGji26czpkd20_buv'
AGENT_INVENTORY = os.path.join(WORKSPACE, 'inventory_agent.json')
SYNC_LOG = os.path.join(WORKSPACE, 'sync_log.json')

def get_credentials():
    """Get or refresh Google OAuth credentials."""
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, 'w') as token:
            token.write(creds.to_json())
    return creds

def get_latest_inventory_file(service):
    """Find the most recently modified Excel file in the inventory folder."""
    results = service.files().list(
        q=f"'{INVENTORY_FOLDER_ID}' in parents and trashed=false and (mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or mimeType='application/vnd.google-apps.spreadsheet')",
        pageSize=10,
        orderBy="modifiedTime desc",
        fields="files(id, name, modifiedTime, size)"
    ).execute()
    files = results.get('files', [])
    if not files:
        return None
    return files[0]  # Most recent

def download_file(service, file_id, file_name):
    """Download file from Google Drive."""
    request = service.files().get_media(fileId=file_id)
    download_path = os.path.join(WORKSPACE, f'download_{file_name}')
    with open(download_path, 'wb') as f:
        downloader = request.execute()
        f.write(downloader)
    return download_path

def convert_excel_to_agent_json(excel_path):
    """Convert zobaze Excel export to agent inventory JSON."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Detect header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    # Map columns
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            h_lower = str(h).lower().strip()
            if 'category' in h_lower: col_map['category'] = i
            elif 'item' in h_lower and 'name' in h_lower: col_map['name'] = i
            elif 'item' in h_lower and 'type' in h_lower: col_map['item_type'] = i
            elif 'variant' in h_lower: col_map['variant'] = i
            elif 'price' in h_lower and 'cost' not in h_lower: col_map['price'] = i
            elif 'cost' in h_lower: col_map['cost'] = i
            elif 'stock' in h_lower: col_map['stock'] = i
            elif 'barcode' in h_lower: col_map['barcode'] = i
            elif 'sku' in h_lower: col_map['sku'] = i
    
    inventory = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row or not row[0]:
            continue
        
        name_idx = col_map.get('name', 2)
        name = row[name_idx] if name_idx < len(row) else None
        if not name or not str(name).strip():
            continue
        
        stock_idx = col_map.get('stock', 6)
        stock = row[stock_idx] if stock_idx < len(row) else 0
        
        price_idx = col_map.get('price', 4)
        price = row[price_idx] if price_idx < len(row) else 0
        
        cost_idx = col_map.get('cost', 5)
        cost = row[cost_idx] if cost_idx < len(row) else 0
        
        inventory.append({
            'category': row[col_map.get('category', 0)] if col_map.get('category', 0) < len(row) else '',
            'name': str(name).strip(),
            'variant': str(row[col_map.get('variant', 3)] if col_map.get('variant', 3) < len(row) else ''),
            'price': float(price) if price else 0,
            'cost': float(cost) if cost else 0,
            'stock': int(stock) if stock else 0,
            'barcode': str(row[col_map.get('barcode', 7)] if col_map.get('barcode', 7) < len(row) else ''),
            'in_stock': bool(stock and int(stock) > 0)
        })
    
    return inventory

def log_sync(file_name, file_time, item_count, in_stock_count):
    """Log sync event."""
    log_entry = {
        'timestamp': datetime.now().isoformat(),
        'file_name': file_name,
        'file_modified': file_time,
        'items_synced': item_count,
        'in_stock': in_stock_count
    }
    
    logs = []
    if os.path.exists(SYNC_LOG):
        with open(SYNC_LOG, 'r') as f:
            logs = json.load(f)
    logs.append(log_entry)
    # Keep last 100 entries
    logs = logs[-100:]
    with open(SYNC_LOG, 'w') as f:
        json.dump(logs, f, indent=2)

def main():
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Starting inventory sync...")
    
    try:
        creds = get_credentials()
        service = build('drive', 'v3', credentials=creds)
        
        # Get latest file
        latest = get_latest_inventory_file(service)
        if not latest:
            print("No inventory files found in Drive folder.")
            return
        
        print(f"Found: {latest['name']} (modified: {latest['modifiedTime']})")
        
        # Check if we already have this version
        if os.path.exists(SYNC_LOG):
            with open(SYNC_LOG, 'r') as f:
                logs = json.load(f)
            for entry in reversed(logs):
                if entry.get('file_modified') == latest['modifiedTime']:
                    print("Already synced this version. Skipping.")
                    return
        
        # Download
        excel_path = download_file(service, latest['id'], latest['name'])
        print(f"Downloaded to {excel_path}")
        
        # Convert
        inventory = convert_excel_to_agent_json(excel_path)
        in_stock = sum(1 for i in inventory if i['in_stock'])
        
        # Save
        with open(AGENT_INVENTORY, 'w') as f:
            json.dump(inventory, f, indent=2)
        
        # Cleanup download
        os.remove(excel_path)
        
        # Log
        log_sync(latest['name'], latest['modifiedTime'], len(inventory), in_stock)
        
        print(f"✅ Sync complete: {len(inventory)} items ({in_stock} in stock)")
        
    except Exception as e:
        print(f"❌ Sync failed: {e}")
        raise

if __name__ == '__main__':
    main()
